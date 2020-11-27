import openpyxl


class OperationExcel:

    def __init__(self, file_path=None, file_name=None, sheetname=None):
        if file_path is None:
            file_path = "D:\\code\\python\\AutoApiTestStudy\\excel_file\\"
        if file_name is None:
            file_name = "api_test"
        self.excel_file = file_path + file_name + '.xlsx'
        self.wb = openpyxl.load_workbook(self.excel_file)
        # self.sheets = self.wb.sheetnames  # 获取表格所有的sheet 名称
        if sheetname is None:
            self.sheet = self.wb.worksheets[0]
        else:
            self.sheet = self.wb[sheetname]
        self.max_row = self.sheet.max_row
        self.max_column = self.sheet.max_column

    def get_cell_value(self, row, column):
        row, column = int(row), int(column)
        if row > self.max_row:
            return 'error, row is out of limit'
        if column > self.max_column:
            return 'error, column is out of limit'
        return self.sheet.cell(row, column).value

    def save_to_excel_by_cell(self, row, column, value):
        row, column = int(row), int(column)
        self.sheet.cell(row, column).value = value
        self.wb.save(self.excel_file)

    def get_column_value(self, column_name):
        column = self.sheet[column_name]
        column_data = {}

        for i in range(1, self.max_row):
            key = str(column[i].value)
            column_data[key] = i + 1

        return column_data


if __name__ == '__main__':
    test = OperationExcel()
    print(test.get_column_value('A'))

    # test.save_to_excel_by_cell(2, 2, 'test1')
    # test.save_to_excel_by_cell(3, 3, 'test2')
    # print(test.get_cell_value(6, 6))
    # print(test.get_cell_value(6, 7))
